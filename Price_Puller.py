import requests
import os
import logging
import sys
import time
import random
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from urllib3.util import Retry
from requests.adapters import HTTPAdapter

# Logging
logging.basicConfig(
    filename='price_puller.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Rotating User-Agents
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
]

def load_urls(file_path):
    urls = {}
    try:
        with open(file_path, 'r') as f:
            for line in f:
                original = line.strip()
                if not original or original.startswith('#') or ',' not in original:
                    continue
                parts = original.split(',', 1)
                sheet, url = parts[0].strip(), parts[1].strip()
                if sheet and url:
                    urls[sheet] = url
        print(f"✅ Loaded {len(urls)} URL(s)")
    except Exception as e:
        print(f"🚨 Error reading {file_path}: {e}")
        logging.exception(f"Error reading {file_path}")
    return urls

def print_progress_bar(iteration, total, prefix='', suffix='', length=50, fill='█', start_time=None):
    percent = f"{100 * (iteration / float(total)):.1f}"
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)

    eta_display = ""
    if start_time and iteration > 0:
        elapsed = time.time() - start_time
        eta = (elapsed / iteration) * (total - iteration)
        eta_display = f" | ETA: {time.strftime('%M:%S', time.gmtime(eta))}"

    sys.stdout.write(f'\r{prefix} |{bar}| {percent}% {suffix}{eta_display}')
    sys.stdout.flush()

def is_html_response(text):
    return '<html' in text.lower()

def get_avg_price(url, sheet_name, output_file='Price_Puller.xlsx'):
    print(f"\n📥 Processing: {sheet_name}")

    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Referer": "https://www.autotrader.com/",
        "X-Requested-With": "XMLHttpRequest"
    }

    session = requests.Session()
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    session.mount('https://', HTTPAdapter(max_retries=retries))

    try:
        print(f"🌐 Connecting to: {url}")
        response = session.get(url, headers=headers, timeout=10)
        print(f"🔁 HTTP status: {response.status_code}")

        if is_html_response(response.text):
            print(f"🚫 Blocked or bad response for {sheet_name} (HTML content detected).")
            return

        if response.status_code == 403:
            print(f"❌ 403 Forbidden for {sheet_name}")
            return
        elif response.status_code != 200:
            print(f"❌ HTTP {response.status_code}: {response.text[:100]}")
            return

        try:
            data = response.json()
        except Exception as e:
            print(f"🚨 JSON parse error: {e}")
            return

        links = data.get('links', [])
        if not links:
            print(f"⚠️ No price data for {sheet_name}")
            return

        today = date.today()
        today_str = today.strftime('%d%b%Y').upper()
        row_data = {'Date': today_str}
        price_years = []

        for link in links:
            year = link.get('value')
            price = link.get('avgPrice')
            if year and price:
                year = int(year)
                row_data[year] = price
                price_years.append(year)

        print(f"📊 Years found: {price_years}")

        anchor = datetime.strptime("22Apr2025", "%d%b%Y").date()
        row_num = (today - anchor).days + 2

        if os.path.exists(output_file):
            wb = load_workbook(output_file)
        else:
            wb = Workbook()

        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        headers = [cell.value for cell in ws[1] if cell.value]

        if not headers:
            headers = ['Date'] + sorted(price_years)
            for i, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=i)
                cell.value = h
                cell.font = Font(bold=True)
        else:
            missing = [y for y in price_years if y not in headers]
            for y in sorted(missing):
                ws.cell(row=1, column=ws.max_column + 1, value=y).font = Font(bold=True)
            headers = [cell.value for cell in ws[1] if cell.value]

        col_map = {header: i + 1 for i, header in enumerate(headers)}
        ws.cell(row=row_num, column=col_map['Date'], value=today_str)

        for year in price_years:
            col = col_map.get(year)
            if col:
                val = round(row_data[year], 2)
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.number_format = '"$"#,##0.00'

        wb.save(output_file)
        print(f"✅ Saved data for {sheet_name}")

    except requests.exceptions.Timeout:
        print(f"⏰ Timeout while connecting to {sheet_name}")
    except requests.RequestException as e:
        print(f"🌐 Network error: {e}")
    except Exception as e:
        print(f"🔥 Unexpected error in {sheet_name}: {e}")

def main():
    url_file = 'car_urls.txt'
    print("🚀 Starting script...\n")

    if not os.path.exists(url_file):
        print(f"❌ File not found: {url_file}")
        return

    urls = load_urls(url_file)
    if not urls:
        print("🚫 No URLs found.")
        return

    total = len(urls)
    start_time = time.time()
    print_progress_bar(0, total, prefix='Progress', suffix='Complete', length=50, start_time=start_time)

    for count, (sheet_name, url) in enumerate(urls.items(), 1):
        get_avg_price(url, sheet_name)
        time.sleep(random.uniform(2, 4))  # Add delay between requests
        print_progress_bar(count, total, prefix='Progress', suffix='Complete', length=50, start_time=start_time)

    print("\n🎉 Done!")

if __name__ == '__main__':
    main()
